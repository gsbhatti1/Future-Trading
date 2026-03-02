import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os, importlib.util

def load_config():
    spec = importlib.util.spec_from_file_location("config",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "0_config.py"))
    cfg = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cfg)
    return cfg

cfg = load_config()

def to_stars(v):
    if pd.isna(v): return ""
    n = min(5, max(1, round(float(v))))
    return {5:"5/5",4:"4/5",3:"3/5",2:"2/5",1:"1/5"}[n]

STAR_COLORS = {5:"FF00BB00",4:"FF66BB00",3:"FFFFCC00",2:"FFFF8800",1:"FFFF2200"}
HDR_BG = "FF1A1A2E"
HDR_FG = "FFE0E7FF"

df = pd.read_csv(os.path.join(cfg.RESULTS_DIR, "master_results.csv"))
df["RATING"] = df["avg_stars"].apply(to_stars)
front = ["strategy","timeframe","RATING","avg_stars","bot_ready","trade_type"]
rest = [c for c in df.columns if c not in front]
df = df[front + rest]

out = os.path.join(cfg.RESULTS_DIR, "master_results_COOL.xlsx")
df.to_excel(out, index=False)

wb = load_workbook(out)
ws = wb.active
ws.title = "Strategy Ranker"
ws.freeze_panes = "A2"
cols = list(df.columns)

for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor=HDR_BG)
    cell.font = Font(color=HDR_FG, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in range(2, ws.max_row + 1):
    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row, ci)
        v = cell.value
        if cn == "RATING":
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal="center")
        elif cn == "avg_stars" and v:
            s = min(5, max(1, round(float(v))))
            cell.fill = PatternFill("solid", fgColor=STAR_COLORS[s])
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")
        elif cn == "bot_ready":
            if v == "YES":
                cell.fill = PatternFill("solid", fgColor="FF00FF99")
                cell.font = Font(bold=True, color="FF003300")
            else:
                cell.fill = PatternFill("solid", fgColor="FFFF4444")
                cell.font = Font(bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center")
        elif cn == "trade_type":
            if v == "SCALP":
                cell.fill = PatternFill("solid", fgColor="FF00BBFF")
                cell.font = Font(bold=True, color="FF000033")
            elif v == "SWING":
                cell.fill = PatternFill("solid", fgColor="FFBB00FF")
                cell.font = Font(bold=True, color="FFFFFFFF")
            elif v == "INTRADAY":
                cell.fill = PatternFill("solid", fgColor="FFFFAA00")
                cell.font = Font(bold=True, color="FF333300")
            cell.alignment = Alignment(horizontal="center")
        elif "win_rate" in cn and v:
            try:
                wr = float(v)
                if wr >= 80: cell.fill = PatternFill("solid", fgColor="FF00FF66")
                elif wr >= 60: cell.fill = PatternFill("solid", fgColor="FFCCFF99")
                elif wr < 40: cell.fill = PatternFill("solid", fgColor="FFFFCCCC")
                cell.alignment = Alignment(horizontal="center")
            except: pass

for ci, cn in enumerate(cols, 1):
    if "strategy" in cn: w = 55
    elif "timeframe" in cn: w = 12
    elif "RATING" in cn: w = 14
    elif "bot" in cn: w = 12
    elif "trade" in cn: w = 13
    elif "avg" in cn: w = 12
    else: w = 16
    ws.column_dimensions[get_column_letter(ci)].width = w

for i in range(1, ws.max_row + 1):
    ws.row_dimensions[i].height = 22

wb.save(out)
print(f"COOL Excel saved -> {out}")
