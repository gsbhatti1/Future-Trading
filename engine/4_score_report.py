# FULL NEW CONTENT FOR 4_score_report.py

import os
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "db" / "strategies.db"
RESULTS_CSV = ROOT / "results" / "raw_results.csv"
REPORTS_DIR = ROOT / "reports"
STRATEGIES_DIR = ROOT / "strategies"

GEM_WIN_RATE = 70.0
TRASH_WIN_RATE = 50.0
MAX_DD_LIMIT = None  # placeholder, we don't have DD yet

STATUS_GEM = "GEM"
STATUS_PASS = "PASS"
STATUS_TRASH = "TRASH"


def ensure_dirs() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    (STRATEGIES_DIR / "best").mkdir(parents=True, exist_ok=True)
    (STRATEGIES_DIR / "staging").mkdir(parents=True, exist_ok=True)
    (STRATEGIES_DIR / "archive" / "trash").mkdir(parents=True, exist_ok=True)


def load_results_from_db() -> pd.DataFrame:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DB not found: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql(
            """
            SELECT
                s.strategy_name,
                br.symbol AS ticker,
                br.timeframe,
                br.total_trades,
                br.win_rate,
                br.total_return,
                br.profit_factor,
                br.status,
                br.composite_score,
                br.id AS backtest_id
            FROM backtest_results br
            JOIN strategies s ON s.id = br.strategy_id
            """,
            conn,
        )
    finally:
        conn.close()
    return df


def classify_row(row: pd.Series) -> str:
    win_rate = float(row.get("win_rate") or 0.0)
    if win_rate >= GEM_WIN_RATE:
        return STATUS_GEM
    if win_rate < TRASH_WIN_RATE:
        return STATUS_TRASH
    return STATUS_PASS


def compute_composite_score(row: pd.Series) -> float:
    win_rate = float(row.get("win_rate") or 0.0)
    ret = float(row.get("total_return") or 0.0)
    trades = float(row.get("total_trades") or 0.0)

    score = win_rate * 0.5 + ret * 0.4 + trades * 0.1
    return score


def apply_scoring(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["status"] = df.apply(classify_row, axis=1)
    df["composite_score"] = df.apply(compute_composite_score, axis=1)
    df = df.sort_values("composite_score", ascending=False)
    return df


def update_db_scores(df: pd.DataFrame) -> None:
    if df.empty:
        return
    conn = sqlite3.connect(DB_PATH)
    try:
        cur = conn.cursor()
        now = datetime.utcnow().isoformat()
        for _, row in df.iterrows():
            cur.execute(
                """
                UPDATE backtest_results
                SET status = ?, composite_score = ?, run_timestamp = ?
                WHERE id = ?
                """,
                (
                    row["status"],
                    float(row["composite_score"]),
                    now,
                    int(row["backtest_id"]),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def move_strategy_files(df: pd.DataFrame) -> None:
    # We don't know real source filenames; we assume one file per strategy in staging
    if df.empty:
        return
    for strat in df["strategy_name"].unique():
        status = df[df["strategy_name"] == strat]["status"].iloc[0]
        src_py = STRATEGIES_DIR / "staging" / f"{strat}.py"
        if not src_py.exists():
            continue
        if status == STATUS_GEM:
            dst = STRATEGIES_DIR / "best" / src_py.name
        elif status == STATUS_TRASH:
            dst = STRATEGIES_DIR / "archive" / "trash" / src_py.name
        else:
            continue
        dst.parent.mkdir(parents=True, exist_ok=True)
        src_py.replace(dst)


def generate_excel(df: pd.DataFrame) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = REPORTS_DIR / f"strategy_rankings_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rankings"

    cols = [
        "strategy_name",
        "ticker",
        "timeframe",
        "total_trades",
        "win_rate",
        "total_return",
        "profit_factor",
        "status",
        "composite_score",
    ]
    df_out = df[cols].copy()

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    status_col_idx = cols.index("status") + 1
    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx in range(2, max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col_idx).value
        if status_val == STATUS_GEM:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status_val == STATUS_TRASH:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(out_path)
    return out_path


def main() -> None:
    ensure_dirs()
    df = load_results_from_db()
    if df.empty:
        print("No backtest results in DB.")
        return
    df_scored = apply_scoring(df)
    update_db_scores(df_scored)
    move_strategy_files(df_scored)
    report_path = generate_excel(df_scored)
    gem_count = int((df_scored["status"] == STATUS_GEM).sum())
    print(f"GEMs: {gem_count}")
    print(f"Excel report: {report_path}")


if __name__ == "__main__":
    main()
